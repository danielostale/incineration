# Bloque 1 — imports and project paths
from __future__ import annotations

from pathlib import Path
import io
import pandas as pd


# Project-level data directory.
# This allows run.py to call write_model_excel_report(df, excel_path)
# without separately passing input paths.
DATA_DIR = Path(__file__).resolve().parents[1] / "data"

DEFAULT_EXOGENOUS_PATH = DATA_DIR / "timeseries_exog.csv"
DEFAULT_PARAMS_PATH = DATA_DIR / "params_scalar.csv"
DEFAULT_AIR_PATH = DATA_DIR / "air_pollutant_factors.csv"
DEFAULT_INITIAL_PATH = DATA_DIR / "initial_conditions.csv"


# Bloque 2 — CSV utilities and required input resolution
def _read_csv_auto(path: Path | str) -> pd.DataFrame:
    """Read CSV files saved with either comma or semicolon separators."""
    path = Path(path)
    raw = path.read_text(encoding="utf-8-sig")
    sep = ";" if ";" in raw.split("\n", 1)[0] else ","
    return pd.read_csv(io.StringIO(raw), sep=sep)


def _existing_path(*candidates: str | Path | None) -> Path | None:
    """Return the first existing path among candidate paths."""
    for candidate in candidates:
        if candidate is None:
            continue
        candidate_path = Path(candidate)
        if candidate_path.exists():
            return candidate_path
    return None


def _required_input_path(label: str, *candidates: str | Path | None) -> Path:
    """
    Resolve a required input file.

    The documentation workbook is not allowed to silently omit input data.
    If a required input CSV cannot be found, the report fails clearly.
    """
    path = _existing_path(*candidates)
    if path is None:
        checked = [str(c) for c in candidates if c is not None]
        raise FileNotFoundError(
            "Cannot build MSW_Model_Documentation.xlsx because required input "
            f"'{label}' was not found. Checked: {checked}"
        )
    return path


# Bloque 3 — master variable catalogue
# This block contains the authoritative manual dictionary of model variables.
# It should eventually document every input, output, state, parameter, transfer,
# diagnostic and environmental variable used by the model.

MASTER_COLUMNS = [
    "variable",
    "symbol",
    "input_output",
    "type",
    "subtype",
    "unit",
    "time_index",
    "equation",
    "definition",
    "model_block",
    "code_module",
    "source_status",
    "source",
    "used_in",
    "computed_from",
    "interpretation",
    "critical_note",
    "appears_in_sheet",
]


def _row(
    variable: str,
    symbol: str = "",
    input_output: str = "",
    type: str = "",
    subtype: str = "",
    unit: str = "",
    time_index: str = "",
    equation: str = "",
    definition: str = "",
    model_block: str = "",
    code_module: str = "",
    source_status: str = "",
    source: str = "",
    used_in: str = "",
    computed_from: str = "",
    interpretation: str = "",
    critical_note: str = "",
    appears_in_sheet: str = "",
) -> dict:
    return {
        "variable": variable,
        "symbol": symbol,
        "input_output": input_output,
        "type": type,
        "subtype": subtype,
        "unit": unit,
        "time_index": time_index,
        "equation": equation,
        "definition": definition,
        "model_block": model_block,
        "code_module": code_module,
        "source_status": source_status,
        "source": source,
        "used_in": used_in,
        "computed_from": computed_from,
        "interpretation": interpretation,
        "critical_note": critical_note,
        "appears_in_sheet": appears_in_sheet,
    }


def _manual_variable_catalogue() -> list[dict]:
    """
    Reviewer-facing dictionary of all central model variables.

    This is intentionally manual. The model is small enough that the workbook
    should not hide important variables behind automatic labels such as
    "computed by model". Any variable not listed here is still appended by
    build_all_variables(), but it is flagged for documentation.
    """
    rows: list[dict] = []

    # ------------------------------------------------------------------
    # Scalar parameters: generation and behavior
    # ------------------------------------------------------------------
    rows += [
        _row("gbar_H", r"\\bar g_H", "Input", "Parameter", "Generation intensity", "ton/served-capita/yr", "fixed", "eq. 3", "Baseline per-capita household MSW generation before awareness-driven reduction.", "Generator behavior", "agents/generator_base.py; agents/households.py", "assumption", "params_scalar.csv", "gen_total; Q_HAUL", "Provided by modeller", "Higher values mechanically increase household waste generation.", "Needs local observed benchmark or calibration; not evidence by itself.", "Parameters_Assumptions"),
        _row("gbar_SC", r"\\bar g_{SC}", "Input", "Parameter", "Generation intensity", "ton/served-capita/yr", "fixed", "eq. 3 / Stage-0 simplification", "Small-commercial generation intensity scaled to served population in the current implementation.", "Generator behavior", "agents/generator_base.py; agents/small_commercial.py", "assumption", "params_scalar.csv", "gen_total; Q_HAUL", "Provided by modeller", "Higher values increase small-commercial generation.", "Conceptually weak if SC should scale with activity rather than population; acceptable only as Stage-0 simplification.", "Parameters_Assumptions"),
        _row("gbar_LG", r"\\bar g_{LG}", "Input", "Parameter", "Generation intensity", "ton/served-capita/yr", "fixed", "eq. 3 / Stage-0 simplification", "Large-generator generation intensity scaled to served population in the current implementation.", "Generator behavior", "agents/generator_base.py; agents/large_generators.py", "assumption", "params_scalar.csv", "gen_total; Q_HAUL", "Provided by modeller", "Higher values increase large-generator generation.", "Conceptually weak if LG should scale with institutional/commercial activity; flag for later replacement.", "Parameters_Assumptions"),
        _row("rho_min_H", r"\\rho^{min}_H", "Input", "Parameter", "Behavioral floor", "ratio", "fixed", "eq. 3", "Irreducible household generation floor under maximum awareness.", "Generator behavior", "forms.py; agents/generator_base.py", "assumption", "params_scalar.csv", "gen_total", "Provided by modeller", "Lower values imply education can reduce more waste generation.", "Strong behavioral assumption; sensitivity analysis required.", "Parameters_Assumptions"),
        _row("rho_min_SC", r"\\rho^{min}_{SC}", "Input", "Parameter", "Behavioral floor", "ratio", "fixed", "eq. 3", "Irreducible small-commercial generation floor under maximum awareness.", "Generator behavior", "forms.py; agents/generator_base.py", "assumption", "params_scalar.csv", "gen_total", "Provided by modeller", "Lower values imply stronger prevention response among small-commercial generators.", "Requires justification; firms may respond differently from households.", "Parameters_Assumptions"),
        _row("rho_min_LG", r"\\rho^{min}_{LG}", "Input", "Parameter", "Behavioral floor", "ratio", "fixed", "eq. 3", "Irreducible large-generator generation floor under maximum awareness.", "Generator behavior", "forms.py; agents/generator_base.py", "assumption", "params_scalar.csv", "gen_total", "Provided by modeller", "Lower values imply stronger prevention response among large generators.", "Potential inconsistency if LG prevention should be contract-price driven rather than education-driven.", "Parameters_Assumptions"),
        _row("kappa_rho", r"\\kappa_\\rho", "Input", "Parameter", "Behavioral shape", "1/awareness-unit", "fixed", "eq. 3 / forms.rho", "Speed at which awareness reduces generation toward the irreducible floor.", "Behavioral closed form", "forms.py", "assumption", "params_scalar.csv", "gen_total", r"rho_min + (1-rho_min)exp(-kappa_rho A)", "Higher values make education reduce generation faster.", "Reviewer-sensitive; quantitative results can depend heavily on this functional form.", "Parameters_Assumptions"),
        _row("pbar", r"\\bar p", "Input", "Inactive legacy parameter", "Formal participation", "$/ton", "fixed", "removed / former eq. 4", "Political threshold above which a general-waste fee would reduce formal participation in earlier model versions.", "Formal participation", "not active in current code", "assumption", "params_scalar.csv", "none unless leakage/theta reintroduced", "Provided by modeller", "Should not affect current results.", "Current implementation assumes full formal participation; keeping this without inactive flag is misleading.", "Parameters_Assumptions"),
        _row("kappa_theta", r"\\kappa_\\theta", "Input", "Inactive legacy parameter", "Formal participation", "1/($/ton)", "fixed", "removed / former eq. 4", "Sensitivity of formal participation to the general-waste fee above pbar in earlier model versions.", "Formal participation", "not active in current code", "assumption", "params_scalar.csv", "none unless leakage/theta reintroduced", "Provided by modeller", "Should not affect current results.", "Current code removes theta; mark inactive or remove from active inputs.", "Parameters_Assumptions"),
        _row("theta_floor", r"\\theta^{min}", "Input", "Inactive legacy parameter", "Formal participation", "ratio", "fixed", "removed / former eq. 4", "Minimum formal-participation rate at very high general-waste fees in earlier model versions.", "Formal participation", "not active in current code", "assumption", "params_scalar.csv", "none unless leakage/theta reintroduced", "Provided by modeller", "Should not affect current results.", "Conflicts with current no-leakage implementation unless reactivated explicitly.", "Parameters_Assumptions"),
        _row("leakage_H", "", "Input", "Inactive legacy switch", "Leakage", "{0,1}", "fixed", "removed / no-leakage closure", "Switch indicating whether households can leak waste into informal disposal in an earlier model version.", "Formal participation / leakage", "not active in current code", "model", "params_scalar.csv", "none unless leakage/theta reintroduced", "Provided by modeller", "Should not affect current results.", "Contradicts current implementation where informal waste is always zero.", "Parameters_Assumptions"),
        _row("leakage_SC", "", "Input", "Inactive legacy switch", "Leakage", "{0,1}", "fixed", "removed / no-leakage closure", "Switch indicating whether small-commercial entities can leak waste into informal disposal in an earlier model version.", "Formal participation / leakage", "not active in current code", "model", "params_scalar.csv", "none unless leakage/theta reintroduced", "Provided by modeller", "Should not affect current results.", "Contradicts current implementation where informal waste is always zero.", "Parameters_Assumptions"),
        _row("leakage_LG", "", "Input", "Inactive legacy switch", "Leakage", "{0,1}", "fixed", "removed / no-leakage closure", "Switch indicating whether large generators can leak waste into informal disposal in an earlier model version.", "Formal participation / leakage", "not active in current code", "model", "params_scalar.csv", "none unless leakage/theta reintroduced", "Provided by modeller", "Should not affect current results.", "Consistent with no LG leakage but redundant under current no-leakage implementation.", "Parameters_Assumptions"),
    ]

    # Source separation and stocks
    rows += [
        _row("sR0", r"s^0_R", "Input", "Parameter", "Sorting baseline", "share", "fixed", "eq. 8", "Baseline recyclable sorting share before awareness, convenience and price-wedge effects.", "Source separation", "forms.py", "assumption", "params_scalar.csv", "sR; Q_HAUL_R; Q_MRF", "Provided by modeller", "Higher values shift waste to recycling before policy effects.", "Should be calibrated to observed recycling stream composition if possible.", "Parameters_Assumptions"),
        _row("sO0", r"s^0_O", "Input", "Parameter", "Sorting baseline", "share", "fixed", "eq. 8", "Baseline organic sorting share before awareness, convenience and price-wedge effects.", "Source separation", "forms.py", "assumption", "params_scalar.csv", "sO; Q_HAUL_O; Q_CMP", "Provided by modeller", "Higher values shift waste to organics before policy effects.", "Important for organics scenarios; should be linked to waste composition and existing diversion.", "Parameters_Assumptions"),
        _row("aA_R", r"a^A_R", "Input", "Parameter", "Awareness response", "share/awareness-unit", "fixed", "eq. 8", "Effect of normalized awareness on recyclable sorting share.", "Source separation", "forms.py", "assumption", "params_scalar.csv", "sR; Q_HAUL_R", "Provided by modeller", "Higher values make education increase recycling faster.", "Pure assumption unless estimated; use sensitivity analysis.", "Parameters_Assumptions"),
        _row("aA_O", r"a^A_O", "Input", "Parameter", "Awareness response", "share/awareness-unit", "fixed", "eq. 8", "Effect of normalized awareness on organic sorting share.", "Source separation", "forms.py", "assumption", "params_scalar.csv", "sO; Q_HAUL_O", "Provided by modeller", "Higher values make education increase organics sorting faster.", "Pure assumption unless estimated; organics response may differ from recycling response.", "Parameters_Assumptions"),
        _row("aw_R", r"a^w_R", "Input", "Parameter", "Price-wedge response", "share/($/ton)", "fixed", "eq. 8", "Effect of the general-minus-recycling fee wedge on recyclable sorting share.", "Source separation", "forms.py", "assumption", "params_scalar.csv", "sR; Q_HAUL_R", "Provided by modeller", "Higher values make differential pricing shift more waste to recycling.", "Requires behavioral elasticity evidence or robustness checks.", "Parameters_Assumptions"),
        _row("aw_O", r"a^w_O", "Input", "Parameter", "Price-wedge response", "share/($/ton)", "fixed", "eq. 8", "Effect of the general-minus-organics fee wedge on organic sorting share.", "Source separation", "forms.py", "assumption", "params_scalar.csv", "sO; Q_HAUL_O", "Provided by modeller", "Higher values make differential pricing shift more waste to organics.", "Requires behavioral elasticity evidence or robustness checks.", "Parameters_Assumptions"),
        _row("sR_max", r"\\bar s_R", "Input", "Parameter", "Sorting cap", "share", "fixed", "eq. 8", "Upper bound on recyclable sorting share.", "Source separation", "forms.py", "assumption", "params_scalar.csv", "sR", "Provided by modeller", "Caps the maximum possible recycling share.", "Must be consistent with physical recyclable fraction of MSW.", "Parameters_Assumptions"),
        _row("sO_max", r"\\bar s_O", "Input", "Parameter", "Sorting cap", "share", "fixed", "eq. 8", "Upper bound on organic sorting share.", "Source separation", "forms.py", "assumption", "params_scalar.csv", "sO", "Provided by modeller", "Caps the maximum possible organics sorting share.", "Must be consistent with organic fraction of MSW.", "Parameters_Assumptions"),
        _row("gamma", r"\\gamma", "Input", "Parameter", "Stock transition", "awareness-unit/$", "fixed", "eq. 10a", "Marginal effect of municipal education expenditure on awareness stock.", "Municipality behavior", "agents/municipality.py", "calibrate", "params_scalar.csv", "A; gen_total; sorting shares", "Provided by modeller", "Higher values make education reduce generation and increase sorting faster.", "Needs calibration; otherwise education-spending experiments are arbitrary in magnitude.", "Parameters_Assumptions"),
        _row("delta_psi", r"\\delta_\\Psi", "Input", "Parameter", "Stock transition", "convenience-unit/$", "fixed", "eq. 10b", "Marginal effect of municipal infrastructure expenditure on convenience stock.", "Municipality behavior", "agents/municipality.py", "calibrate", "params_scalar.csv", "Psi; sorting shares", "Provided by modeller", "Higher values make X_MUN increase sorting faster when use_convenience=1.", "Needs calibration or scenario interpretation; otherwise X_MUN effects are arbitrary.", "Parameters_Assumptions"),
        _row("use_convenience", "", "Input", "Model switch", "Source separation", "{0,1}", "fixed", "eq. 8 / model switch", "Switch controlling whether convenience stock affects source-separation shares.", "Source separation", "forms.py", "model", "params_scalar.csv", "sR; sO", "Provided by modeller", "Set 1 to activate Psi channel; set 0 to remove it.", "Acceptable as robustness switch; document which results use it.", "Parameters_Assumptions"),
        _row("aPsi_R", r"a^\\Psi_R", "Input", "Parameter", "Convenience response", "share/convenience-unit", "fixed", "eq. 8", "Effect of normalized convenience stock on recyclable sorting share.", "Source separation", "forms.py", "assumption", "params_scalar.csv", "sR; Q_HAUL_R", "Provided by modeller", "Higher values make convenience investments raise recycling faster.", "Requires sensitivity analysis; infrastructure convenience should ideally be spatially grounded.", "Parameters_Assumptions"),
        _row("aPsi_O", r"a^\\Psi_O", "Input", "Parameter", "Convenience response", "share/convenience-unit", "fixed", "eq. 8", "Effect of normalized convenience stock on organic sorting share.", "Source separation", "forms.py", "assumption", "params_scalar.csv", "sO; Q_HAUL_O", "Provided by modeller", "Higher values make convenience investments raise organics sorting faster.", "Requires sensitivity analysis; organics convenience may differ from recycling convenience.", "Parameters_Assumptions"),
        _row("A0", r"A_0", "Input", "Initial condition", "State", "awareness-unit", "0", "initial condition for eq. 10a", "Initial awareness stock at the start of the simulation.", "State initialization", "simulate.py", "assumption", "params_scalar.csv", "A", "Provided by modeller", "Sets the initial level of education/awareness capital.", "If nonzero, needs interpretation relative to prior education history.", "Parameters_Assumptions"),
        _row("Psi0", r"\\Psi_0", "Input", "Initial condition", "State", "convenience-unit", "0", "initial condition for eq. 10b", "Initial convenience stock at the start of the simulation.", "State initialization", "simulate.py", "assumption", "params_scalar.csv", "Psi", "Provided by modeller", "Sets the initial level of convenience infrastructure stock.", "If nonzero, needs interpretation relative to existing infrastructure.", "Parameters_Assumptions"),
    ]

    # Treatment, costs, capital, environment scalar parameters
    rows += [
        _row("eta_R", r"\\eta_R", "Input", "Parameter", "Mechanical recovery", "recovery fraction", "fixed", "eq. 25 / eq. 27", "Mechanical recovery rate of recyclables from mixed general waste sent to separation.", "Treatment / separation", "agents/treatment.py", "assumption", "params_scalar.csv", "Q_MRF; Q_mat; Q_LF", "Provided by modeller", "Higher values increase recovered recyclables from mixed waste.", "Must be justified by separation technology; important for mechanical sorting scenarios.", "Parameters_Assumptions"),
        _row("eta_O", r"\\eta_O", "Input", "Parameter", "Mechanical recovery", "recovery fraction", "fixed", "eq. 25 / eq. 27", "Mechanical recovery rate of organics from mixed general waste sent to separation.", "Treatment / separation", "agents/treatment.py", "assumption", "params_scalar.csv", "Q_CMP; Q_cmp; Q_LF", "Provided by modeller", "Higher values increase recovered organics from mixed waste.", "May be optimistic depending on contamination; needs technology-specific evidence.", "Parameters_Assumptions"),
        _row("c_sep", r"c_{sep}", "Input", "Parameter", "Cost", "$/ton", "fixed", "eq. 28 / eq. 46", "Variable operating cost of mechanical separation per ton processed.", "Treatment / separation", "agents/treatment.py", "assumption", "params_scalar.csv", "net_T", "Provided by modeller", "Higher values worsen economics of mechanical sorting.", "Requires local/technology benchmark; central to whether mixed-waste separation is viable.", "Parameters_Assumptions"),
        _row("sep_share", r"q^{sep share}", "Input", "Scenario variable", "Treatment routing", "share", "fixed/scenario", "eq. 23", "Fraction of mixed general waste mechanically separated before residual treatment.", "Treatment / separation", "agents/treatment.py", "scenario", "params_scalar.csv", "Q_sep; Q_MRF; Q_CMP; Q_LF", "Provided by modeller", "Higher values force more mixed waste through separation.", "In Stage-0 this is exogenous, not optimized; state this clearly.", "Parameters_Assumptions"),
        _row("lambda_MRF", r"\\lambda_{MRF}", "Input", "Parameter", "Reject fraction", "fraction", "fixed", "eq. 32 / eq. 56", "MRF reject fraction: share of MRF throughput not recovered as marketable material.", "Treatment / MRF", "agents/treatment.py; closure.py", "assumption", "params_scalar.csv", "Q_mat; rej_MRF; Q_LF; D_NC", "Provided by modeller", "Higher values reduce material recovery and increase landfill residuals.", "Should reflect contamination and facility performance.", "Parameters_Assumptions"),
        _row("lambda_CMP", r"\\lambda_{CMP}", "Input", "Parameter", "Reject fraction", "fraction", "fixed", "eq. 34 / eq. 56", "Composting reject fraction: share of organics throughput rejected to residual disposal.", "Treatment / composting", "agents/treatment.py; closure.py", "assumption", "params_scalar.csv", "Q_cmp; rej_CMP; Q_LF; D_NC", "Provided by modeller", "Higher values reduce compost output and increase landfill residuals.", "Should reflect contamination and processing technology.", "Parameters_Assumptions"),
        _row("ell_CMP", r"\\ell_{CMP}", "Input", "Parameter", "Biological loss", "fraction", "fixed", "eq. 34 / eq. 55", "Biological process loss in composting, including moisture and CO2 mass loss.", "Treatment / composting", "agents/treatment.py; closure.py", "assumption", "params_scalar.csv", "Q_loss; Q_cmp", "Provided by modeller", "Higher values reduce marketable compost output while preserving mass closure.", "Important for mass closure; do not treat process loss as recovered compost.", "Parameters_Assumptions"),
        _row("lambda_WTE", r"\\lambda_{WTE}", "Input", "Parameter", "Ash fraction", "fraction", "fixed", "eq. 37 / eq. 40", "WTE ash fraction of combusted residual waste.", "Treatment / WTE", "agents/treatment.py", "assumption", "params_scalar.csv", "Q_ash; Q_LF; D_LF", "Provided by modeller", "Higher values reduce WTE landfill diversion benefit because more ash remains.", "WTE reduces mass but does not eliminate final disposal demand.", "Parameters_Assumptions"),
        _row("e_E", r"e_E", "Input", "Conversion factor", "Energy output", "MWh/ton", "fixed", "eq. 38 / eq. 44 / eq. 68", "Electricity generated per ton of WTE throughput.", "Treatment / WTE", "agents/treatment.py; agents/environment.py", "section_4.2.5", "params_scalar.csv", "E_WTE; rev_E; E_net; GHG_net", "Provided by modeller", "Higher values increase electricity revenue and avoided grid emissions.", "Check unit consistency: 0.31 kWh/kg is roughly 0.31 MWh/metric ton.", "Parameters_Assumptions"),
        _row("e_use_MRF", "", "Input", "Parameter", "Energy use", "MWh/ton", "fixed", "eq. 68", "Energy use per ton of MRF throughput.", "Environmental accounting", "agents/environment.py", "assumption", "params_scalar.csv", "E_net", "Provided by modeller", "Higher values reduce net energy balance.", "Assumed; should be benchmarked if energy results matter.", "Parameters_Assumptions"),
        _row("e_use_CMP", "", "Input", "Parameter", "Energy use", "MWh/ton", "fixed", "eq. 68", "Energy use per ton of composting throughput.", "Environmental accounting", "agents/environment.py", "assumption", "params_scalar.csv", "E_net", "Provided by modeller", "Higher values reduce net energy balance.", "Assumed; needs benchmark if used for energy/GHG claims.", "Parameters_Assumptions"),
        _row("e_use_WTE", "", "Input", "Parameter", "Energy use", "MWh/ton", "fixed", "eq. 68", "Internal energy use per ton of WTE throughput.", "Environmental accounting", "agents/environment.py", "assumption", "params_scalar.csv", "E_net", "Provided by modeller", "Higher values reduce net energy benefit of WTE.", "Should be netted consistently against gross WTE electricity output.", "Parameters_Assumptions"),
        _row("e_use_LF", "", "Input", "Parameter", "Energy use", "MWh/ton", "fixed", "eq. 68", "Energy use per ton of landfill demand.", "Environmental accounting", "agents/environment.py", "assumption", "params_scalar.csv", "E_net", "Provided by modeller", "Higher values reduce net energy balance.", "Likely minor but should be documented.", "Parameters_Assumptions"),
        _row("e_use_sep", "", "Input", "Parameter", "Energy use", "MWh/ton", "fixed", "eq. 68", "Energy use per ton of mechanical separation throughput.", "Environmental accounting", "agents/environment.py", "assumption", "params_scalar.csv", "E_net", "Provided by modeller", "Higher values reduce net energy balance when sep_share > 0.", "Relevant only when mechanical separation is active.", "Parameters_Assumptions"),
        _row("delta_LF", r"\\delta_{LF}", "Input", "Parameter", "Stock degradation", "1/yr", "fixed", "eq. 41", "Landfill stock degradation rate.", "Landfill stock", "agents/treatment.py", "assumption", "params_scalar.csv", "S_LF", "Provided by modeller", "Higher values reduce accumulated landfill stock over time.", "Set to zero if S_LF is interpreted as cumulative deposits over the simulation horizon.", "Parameters_Assumptions"),
        _row("S_LF0", r"S^{LF}_0", "Input", "Initial condition", "Landfill stock", "ton", "0", "initial condition for eq. 41", "Initial cumulative landfilled stock.", "State initialization", "simulate.py", "assumption", "params_scalar.csv", "S_LF", "Provided by modeller", "Sets starting landfill stock.", "If stock matters, use observed historical stock or define as simulation-period stock only.", "Parameters_Assumptions"),
        _row("S_ENV0", r"S^{ENV}_0", "Input", "Initial condition", "Informal environmental stock", "ton", "0", "initial condition for eq. 65", "Initial informal waste stock in the environment.", "State initialization / environment", "simulate.py; agents/environment.py", "assumption", "params_scalar.csv", "S_ENV; D_env", "Provided by modeller", "Sets starting environmental leakage stock.", "Current simulate.py may initialize S_ENV directly; ensure this parameter is used if kept.", "Parameters_Assumptions"),
        _row("c_coll", r"c_{coll}", "Input", "Parameter", "Cost", "$/ton", "fixed", "§5.5.1", "Hauler collection and transport cost per ton collected.", "Collection / haulers", "agents/haulers.py", "assumption", "params_scalar.csv", "net_HAUL", "Provided by modeller", "Higher values reduce hauler net balance and increase real system cost if included.", "Needs benchmark from contracts or local hauling cost data.", "Parameters_Assumptions"),
        _row("c_MRF", r"c_{MRF}", "Input", "Parameter", "Cost", "$/ton", "fixed", "eq. 46", "Variable operating cost per ton of MRF throughput.", "Treatment / MRF", "agents/treatment.py", "assumption", "params_scalar.csv", "net_T", "Provided by modeller", "Higher values reduce treatment operator net balance.", "Needs facility benchmark; central to cost comparison.", "Parameters_Assumptions"),
        _row("c_CMP", r"c_{CMP}", "Input", "Parameter", "Cost", "$/ton", "fixed", "eq. 46", "Variable operating cost per ton of composting throughput.", "Treatment / composting", "agents/treatment.py", "assumption", "params_scalar.csv", "net_T", "Provided by modeller", "Higher values reduce treatment operator net balance.", "Needs benchmark; important for organics diversion scenarios.", "Parameters_Assumptions"),
        _row("c_WTE", r"c_{WTE}", "Input", "Parameter", "Cost", "$/ton", "fixed", "eq. 46", "Variable operating cost per ton of WTE throughput.", "Treatment / WTE", "agents/treatment.py", "assumption", "params_scalar.csv", "net_T", "Provided by modeller", "Higher values reduce WTE private viability.", "Strong assumption; WTE O&M varies substantially by facility and pollution-control requirements.", "Parameters_Assumptions"),
        _row("c_LF", r"c_{LF}", "Input", "Parameter", "Cost", "$/ton", "fixed", "eq. 46", "Variable operating cost per ton of final landfill demand.", "Treatment / landfill", "agents/treatment.py", "assumption", "params_scalar.csv", "net_T", "Provided by modeller", "Higher values reduce landfill route profitability.", "Separate from taxes/tipping fees to avoid transfer-cost confusion.", "Parameters_Assumptions"),
        _row("chi_WTE", r"\\chi_{WTE}", "Input", "Capital parameter", "Unit capital cost", "$/ton/year capacity", "fixed", "eq. 61 / eq. 63", "Unit capital cost of WTE capacity.", "Capacity and investment", "future capital module", "calibrate", "params_scalar.csv", "capital charge if implemented", "Provided by modeller", "Higher values worsen WTE investment economics.", "Important but not currently used in Stage-0 recursion unless investment/capital cost module is activated.", "Parameters_Assumptions"),
        _row("alpha_WTE", r"\\alpha_{WTE}", "Input", "Capital parameter", "Capital recovery", "1/yr", "fixed", "eq. 63", "Capital recovery factor for WTE investment.", "Capacity and investment", "future capital module", "calibrate", "params_scalar.csv", "capital charge if implemented", "Provided by modeller", "Higher values increase annualized capital charge.", "Should reflect financing assumptions and asset life; currently inactive unless capital accounting is implemented.", "Parameters_Assumptions"),
        _row("SCC", "SCC", "Input", "Damage parameter", "Climate valuation", "$/tCO2e", "fixed", "eq. 71", "Social cost of carbon applied to net GHG emissions.", "Environmental damages", "agents/government.py; agents/environment.py", "EPA_2023/scenario", "params_scalar.csv", "D_climate; D_env", "Provided by modeller", "Higher values penalize high-GHG scenarios more strongly.", "Use scenario range; SCC choice is normative and materially affects social cost.", "Parameters_Assumptions"),
        _row("g_coll", r"g_{coll}", "Input", "Emission factor", "GHG", "tCO2e/ton", "fixed", "eq. 66", "GHG emission factor for collection and transport per ton hauled.", "Environmental accounting", "agents/environment.py", "assumption", "params_scalar.csv", "GHG_dir; GHG_net", "Provided by modeller", "Higher values penalize high collection tonnage.", "Should reflect route distance, fleet type and fuel use if possible.", "Parameters_Assumptions"),
        _row("g_sep", r"g_{sep}", "Input", "Emission factor", "GHG", "tCO2e/ton", "fixed", "eq. 66", "GHG emission factor for separation per ton mechanically separated.", "Environmental accounting", "agents/environment.py", "assumption", "params_scalar.csv", "GHG_dir; GHG_net", "Provided by modeller", "Higher values penalize mechanical separation.", "Relevant only when sep_share > 0.", "Parameters_Assumptions"),
        _row("g_MRF", r"g_{MRF}", "Input", "Emission factor", "GHG", "tCO2e/ton", "fixed", "eq. 66", "Direct GHG emission factor for MRF throughput.", "Environmental accounting", "agents/environment.py", "assumption", "params_scalar.csv", "GHG_dir; GHG_net", "Provided by modeller", "Higher values penalize recycling throughput before avoided-material credits.", "Avoided material credit is separate; do not net twice.", "Parameters_Assumptions"),
        _row("g_CMP", r"g_{CMP}", "Input", "Emission factor", "GHG", "tCO2e/ton", "fixed", "eq. 66", "Direct GHG emission factor for composting throughput.", "Environmental accounting", "agents/environment.py", "assumption", "params_scalar.csv", "GHG_dir; GHG_net", "Provided by modeller", "Higher values penalize composting throughput.", "Distinguish composting emissions from avoided emissions credit.", "Parameters_Assumptions"),
        _row("g_WTE", r"g_{WTE}", "Input", "Emission factor", "GHG", "tCO2e/ton", "fixed", "eq. 66", "Direct GHG emission factor for WTE throughput.", "Environmental accounting", "agents/environment.py", "assumption", "params_scalar.csv", "GHG_dir; GHG_net", "Provided by modeller", "Higher values penalize WTE scenarios.", "Key contested parameter; requires facility-specific or literature support.", "Parameters_Assumptions"),
        _row("g_LF", r"g_{LF}", "Input", "Emission factor", "GHG", "tCO2e/ton", "fixed", "eq. 66", "Direct GHG emission factor for final landfill demand.", "Environmental accounting", "agents/environment.py", "assumption", "params_scalar.csv", "GHG_dir; GHG_net", "Provided by modeller", "Higher values penalize landfill-heavy scenarios.", "Specify methane capture assumptions and time horizon.", "Parameters_Assumptions"),
        _row("b_MRF", r"b_{MRF}", "Input", "Avoided-emissions parameter", "GHG credit", "tCO2e/ton", "fixed", "eq. 69", "Avoided emissions credit per ton of recovered material.", "Environmental accounting", "agents/environment.py", "assumption", "params_scalar.csv", "B_env; GHG_net", "Provided by modeller", "Higher values favor recycling scenarios.", "High leverage; must avoid double-counting and should vary by material composition.", "Parameters_Assumptions"),
        _row("b_CMP", r"b_{CMP}", "Input", "Avoided-emissions parameter", "GHG credit", "tCO2e/ton", "fixed", "eq. 69", "Avoided emissions credit per ton of compost output.", "Environmental accounting", "agents/environment.py", "assumption", "params_scalar.csv", "B_env; GHG_net", "Provided by modeller", "Higher values favor composting scenarios.", "Needs basis: fertilizer displacement, soil carbon or methane avoidance are different mechanisms.", "Parameters_Assumptions"),
        _row("g_grid", r"g_{grid}", "Input", "Avoided-emissions parameter", "Grid displacement", "tCO2e/MWh", "fixed", "eq. 69", "Grid emissions intensity displaced by WTE electricity generation.", "Environmental accounting", "agents/environment.py", "assumption", "params_scalar.csv", "B_env; GHG_net", "Provided by modeller", "Higher values favor WTE electricity credit.", "Use Florida marginal emissions, not average grid, if evaluating displacement.", "Parameters_Assumptions"),
        _row("d_ENV", r"d_{ENV}", "Input", "Damage parameter", "Informal waste damage", "$/ton/yr", "fixed", "eq. 73", "Flow damage per ton of informal waste stock in the environment.", "Environmental damages", "agents/environment.py", "assumption", "params_scalar.csv", "D_env", "Provided by modeller", "Higher values penalize informal leakage.", "Currently irrelevant if no informal leakage; if activated, needs strong justification.", "Parameters_Assumptions"),
    ]

    # ------------------------------------------------------------------
    # Annual exogenous / scenario variables
    # ------------------------------------------------------------------
    ts = "Exogenous_Inputs"
    rows += [
        _row("year", "t", "Input", "Time index", "Calendar year", "year", "t", "input index", "Simulation year.", "Timing", "config.py; simulate.py", "observed/index", "timeseries_exog.csv", "all annual variables", "Provided by modeller", "Defines the annual simulation period.", "Must be unique and sorted for clean recursion.", ts),
        _row("N", r"N_t", "Input", "Exogenous driver", "Demography", "persons", "t", "eq. 3", "Population served by the municipal solid-waste system.", "Generator behavior", "config.py; simulate.py; agents/generator_base.py", "observed/scenario", "timeseries_exog.csv", "gen_total", "Provided by modeller", "Higher values scale generation proportionally.", "If constant, model assumes no population growth over horizon.", ts),
        _row("p_G", r"p^G_t", "Input", "Policy variable", "Collection fee", "$/ton", "t", "eq. 8 / eq. 9", "General/residual-waste collection fee faced by generators.", "Source separation / pricing", "forms.py; agents/generator_base.py", "scenario", "timeseries_exog.csv", "sR; sO; gen_to_mun", "Provided by modeller", "Higher p_G increases sorting incentives through fee wedges.", "Not a social cost; it is a price/transfer in the model.", ts),
        _row("p_R", r"p^R_t", "Input", "Policy variable", "Collection fee", "$/ton", "t", "eq. 8 / eq. 9", "Recyclables-stream collection fee.", "Source separation / pricing", "forms.py; agents/generator_base.py", "scenario", "timeseries_exog.csv", "sR; gen_to_mun", "Provided by modeller", "Lower p_R relative to p_G encourages recycling.", "Only meaningful relative to p_G.", ts),
        _row("p_O", r"p^O_t", "Input", "Policy variable", "Collection fee", "$/ton", "t", "eq. 8 / eq. 9", "Organics-stream collection fee.", "Source separation / pricing", "forms.py; agents/generator_base.py", "scenario", "timeseries_exog.csv", "sO; gen_to_mun", "Provided by modeller", "Lower p_O relative to p_G encourages organics sorting.", "Only meaningful relative to p_G.", ts),
        _row("p_HAUL", r"p^{HAUL}_t", "Input", "Contract/payment", "Hauler payment", "$/ton", "t", "eq. 49 / eq. 51", "Municipal contract payment to haulers per ton collected.", "Municipality / haulers", "agents/municipality.py", "scenario/contract", "timeseries_exog.csv", "mun_to_haul; net_MUN; net_HAUL", "Provided by modeller", "Higher values move money from MUN to haulers.", "Internal transfer, not necessarily a real resource cost.", ts),
        _row("E_edu", r"E^{MUN}_t", "Input", "Policy flow", "Education spending", "$/yr", "t", "eq. 10a", "Municipal education expenditure that accumulates into awareness stock A.", "Municipality behavior", "agents/municipality.py", "scenario", "timeseries_exog.csv", "A; gen_total; sR; sO", "Provided by modeller", "Higher values increase future awareness.", "Effect depends entirely on gamma; must be calibrated or sensitivity-tested.", ts),
        _row("X_MUN", r"X^{MUN}_t", "Input", "Policy flow", "Convenience investment", "$/yr", "t", "eq. 10b", "Municipal infrastructure/convenience expenditure that accumulates into Psi.", "Municipality behavior", "agents/municipality.py", "scenario", "timeseries_exog.csv", "Psi; sR; sO", "Provided by modeller", "Higher values increase future convenience stock.", "Effect depends on delta_psi and use_convenience.", ts),
        _row("p_mat", r"p^{mat}_t", "Input", "Market price", "Recovered materials", "$/ton", "t", "eq. 44", "Price received for recovered marketable materials.", "Treatment revenues", "agents/treatment.py", "scenario/market", "timeseries_exog.csv", "rev_mat; net_T", "Provided by modeller", "Higher values improve recycling revenue.", "Should be stochastic or scenario-based; constant price is simplifying.", ts),
        _row("p_cmp", r"p^{cmp}_t", "Input", "Market price", "Compost", "$/ton", "t", "eq. 44", "Price received for marketable compost output.", "Treatment revenues", "agents/treatment.py", "scenario/market", "timeseries_exog.csv", "rev_cmp if recorded; net_T", "Provided by modeller", "Higher values improve composting revenue.", "Zero baseline is defensible only if no reliable compost market/outlet.", ts),
        _row("p_E", r"p^E_t", "Input", "Market price", "Electricity", "$/MWh", "t", "eq. 44", "Electricity price received for WTE electricity generation.", "Treatment revenues", "agents/treatment.py", "scenario/market", "timeseries_exog.csv", "rev_E; net_T", "Provided by modeller", "Higher values improve WTE revenue.", "Should reflect PPA or wholesale-market assumption.", ts),
        _row("tax_LF", r"\\tau^{LF}_t", "Input", "Policy/tax", "Landfill tax", "$/ton", "t", "eq. 54 / eq. 47", "Landfill tax or surcharge per ton of final landfill demand.", "Government / treatment accounting", "agents/government.py", "scenario", "timeseries_exog.csv", "landfill_tax; net_T", "Provided by modeller", "Higher values penalize landfill through operator accounts.", "Internal transfer if GOV inside boundary; do not double-count as social cost.", ts),
        _row("t_MRF", r"t^{MRF}_t", "Input", "Gate fee", "MRF tipping", "$/ton", "t", "eq. 43", "Tipping fee charged for MRF throughput.", "Treatment revenues", "agents/treatment.py", "scenario/contract", "timeseries_exog.csv", "tipping; net_HAUL; net_T", "Provided by modeller", "Higher values shift revenue to treatment operator from haulers.", "Internal transfer; not system cost directly.", ts),
        _row("t_CMP", r"t^{CMP}_t", "Input", "Gate fee", "Composting tipping", "$/ton", "t", "eq. 43", "Tipping fee charged for composting/organics throughput.", "Treatment revenues", "agents/treatment.py", "scenario/contract", "timeseries_exog.csv", "tipping; net_HAUL; net_T", "Provided by modeller", "Higher values shift revenue to treatment operator from haulers.", "Internal transfer; not system cost directly.", ts),
        _row("t_WTE", r"t^{WTE}_t", "Input", "Gate fee", "WTE tipping", "$/ton", "t", "eq. 43", "Tipping fee charged for WTE throughput.", "Treatment revenues", "agents/treatment.py", "scenario/contract", "timeseries_exog.csv", "tipping; net_HAUL; net_T", "Provided by modeller", "Higher values shift revenue to treatment operator from haulers.", "Do not confuse tipping fee with WTE operating cost or environmental damage.", ts),
        _row("t_LF", r"t^{LF}_t", "Input", "Gate fee", "Landfill tipping", "$/ton", "t", "eq. 43", "Tipping fee charged for landfill demand.", "Treatment revenues", "agents/treatment.py", "scenario/contract", "timeseries_exog.csv", "tipping; net_HAUL; net_T", "Provided by modeller", "Higher values shift revenue to treatment operator from haulers.", "Do not confuse landfill tipping fee with landfill tax or real landfill O&M cost.", ts),
        _row("K_MRF", r"K^{MRF}_t", "Input", "Capacity", "MRF", "tons/year", "t", "eq. 31 / eq. 62", "Installed MRF processing capacity available in each year.", "Treatment capacity", "agents/treatment.py", "observed/scenario", "timeseries_exog.csv", "Q_MRF; Q_LF", "Provided by modeller", "Higher capacity allows more recyclable throughput before overflow.", "If constant, assumes no MRF investment/retirement.", ts),
        _row("K_CMP", r"K^{CMP}_t", "Input", "Capacity", "Composting", "tons/year", "t", "eq. 33 / eq. 62", "Installed composting/organics processing capacity available in each year.", "Treatment capacity", "agents/treatment.py", "observed/scenario", "timeseries_exog.csv", "Q_CMP; Q_LF", "Provided by modeller", "Higher capacity allows more organics throughput before overflow.", "Likely binding in organics scenarios; needs local facility evidence.", ts),
        _row("K_WTE", r"K^{WTE}_t", "Input", "Capacity", "Waste-to-energy", "tons/year", "t", "eq. 36 / eq. 62", "Installed WTE processing capacity available in each year.", "Treatment capacity", "agents/treatment.py", "scenario", "timeseries_exog.csv", "Q_WTE; Q_LF; E_WTE; GHG_net", "Provided by modeller", "Higher capacity shifts residual waste away from direct landfill when residual is available.", "Baseline should match actual post-RRF capacity assumption; positive value implies active WTE capacity.", ts),
        _row("K_LF", r"K^{LF}_t", "Input", "Capacity", "Landfill annual flow", "tons/year", "t", "flow capacity / implementation-specific", "Annual landfill throughput capacity, if used as a flow constraint.", "Treatment capacity / landfill", "config.py; future treatment constraints", "scenario", "timeseries_exog.csv", "Q_LF if implemented", "Provided by modeller", "Can restrict annual landfill demand if implemented.", "Conceptual issue: model text distinguishes annual flow capacity from remaining permitted capacity RC_LF.", ts),
        _row("RC_LF", r"RC^{LF}_t", "Input", "Capacity stock", "Remaining landfill capacity", "tons", "t", "eq. 42", "Remaining permitted landfill capacity stock.", "Landfill stock / capacity", "config.py; future stock constraint", "observed/scenario", "timeseries_exog.csv", "landfill feasibility", "Provided by modeller", "Lower values indicate reduced remaining disposal space.", "If constant across years, depletion is not being represented; better as initial stock plus law of motion.", ts),
    ]

    # ------------------------------------------------------------------
    # Air-pollutant factor table variables
    # ------------------------------------------------------------------
    ap = "Air_Pollutant_Factors"
    rows += [
        _row("pollutant", "p", "Input", "Index", "Local air pollutant", "text", "p", "eq. 67 / eq. 72", "Name of the local air pollutant being tracked.", "Environmental externalities", "agents/environment.py", "model/index", "air_pollutant_factors.csv", "D_air", "Provided by modeller", "Defines rows over which local air damages are summed.", "Use consistent pollutant names; PM25 may be code-safe label for PM2.5.", ap),
        _row("a_WTE", r"a_{p,WTE}", "Input", "Emission factor", "Local air pollutant", "kg pollutant/ton WTE", "p", "eq. 67", "Emission factor of pollutant p from WTE treatment.", "Environmental externalities", "agents/environment.py", "assumption", "air_pollutant_factors.csv", "AP_p; D_air", "Provided by modeller", "Higher values increase WTE local air damages.", "Requires facility-specific emissions or defensible literature source for WTE evaluation.", ap),
        _row("a_LF", r"a_{p,LF}", "Input", "Emission factor", "Local air pollutant", "kg pollutant/ton LF", "p", "eq. 67", "Emission factor of pollutant p from landfill disposal.", "Environmental externalities", "agents/environment.py", "assumption", "air_pollutant_factors.csv", "AP_p; D_air", "Provided by modeller", "Higher values increase landfill local air damages.", "Should reflect landfill emissions pathway and whether pollutant is relevant for LF.", ap),
        _row("a_sep", r"a_{p,sep}", "Input", "Emission factor", "Local air pollutant", "kg pollutant/ton separated", "p", "eq. 67", "Emission factor of pollutant p from mechanical separation.", "Environmental externalities", "agents/environment.py", "assumption", "air_pollutant_factors.csv", "AP_p; D_air", "Provided by modeller", "Higher values increase air damages from mechanical separation.", "Usually small, but relevant when sep_share is high.", ap),
        _row("D_p", r"MD_p", "Input", "Damage parameter", "Local air pollutant", "$/kg pollutant", "p", "eq. 72", "Marginal damage of one additional kg of pollutant p emitted.", "Environmental damages", "agents/environment.py", "assumption/damage model", "air_pollutant_factors.csv", "D_air; D_env", "Provided by modeller", "Higher values increase monetized air damages for a given emission factor.", "Highly spatial and exposure-dependent; placeholder values should not support strong policy claims.", ap),
    ]

    # ------------------------------------------------------------------
    # Endogenous outputs from simulate.py
    # ------------------------------------------------------------------
    out = "Endogenous_Outputs"
    rows += [
        _row("A", r"A_t", "Output", "Endogenous state", "Awareness stock", "index", "t", "eq. 10a", "Awareness stock at the beginning of each simulation year.", "Municipality behavior", "simulate.py; agents/municipality.py", "computed", "model output", "rho; sR; sO", "Previous A plus gamma*E_edu", "Higher A reduces generation and increases sorting.", "Do not interpret as observed awareness unless calibrated.", out),
        _row("Psi", r"\\Psi_t", "Output", "Endogenous state", "Convenience stock", "index", "t", "eq. 10b", "Convenience stock at the beginning of each simulation year.", "Municipality behavior", "simulate.py; agents/municipality.py", "computed", "model output", "sR; sO when use_convenience=1", "Previous Psi plus delta_psi*X_MUN", "Higher Psi increases sorting if convenience channel is active.", "Do not interpret as physical infrastructure count unless calibrated.", out),
        _row("sR", r"s_{H,R,t}", "Output", "Endogenous behavioral variable", "Household sorting share", "share", "t", "eq. 8", "Household recyclable sorting share recorded for dashboard/output.", "Source separation", "forms.py; simulate.py", "computed", "model output", "Q_HAUL_R; Q_MRF", "sorting_shares(A,Psi,p_G,p_R,p_O)", "Higher values shift waste toward recycling.", "Currently records household share only; not separate by SC/LG in output.", out),
        _row("sO", r"s_{H,O,t}", "Output", "Endogenous behavioral variable", "Household sorting share", "share", "t", "eq. 8", "Household organic sorting share recorded for dashboard/output.", "Source separation", "forms.py; simulate.py", "computed", "model output", "Q_HAUL_O; Q_CMP", "sorting_shares(A,Psi,p_G,p_R,p_O)", "Higher values shift waste toward organics.", "Currently records household share only; not separate by SC/LG in output.", out),
        _row("sG", r"s_{H,G,t}", "Output", "Endogenous behavioral variable", "Household sorting share", "share", "t", "eq. 8", "Household general/residual sorting share recorded for dashboard/output.", "Source separation", "forms.py; simulate.py", "computed", "model output", "Q_HAUL_G", "1 - sR - sO", "Higher values leave more waste in residual/general stream.", "Currently records household share only; not separate by SC/LG in output.", out),
        _row("gen_total", r"\\sum_g Q_{g,t}", "Output", "Endogenous flow", "Generation", "tons/year", "t", "eq. 3", "Total municipal solid waste generation across all generator types.", "Generator behavior", "simulate.py; agents/generator_base.py", "computed", "model output", "Q_HAUL; mass_residual", "sum over generator outputs", "Main physical scale of the system.", "Falls over time if A increases and rho(A) reduces generation, even with constant population.", out),
        _row("informal", r"\\sum_g Q^I_{g,t}", "Output", "Endogenous flow", "Informal leakage", "tons/year", "t", "eq. 65 / no-leakage closure", "Waste leaking outside the formal system.", "Generator behavior / environment", "simulate.py", "computed", "model output", "S_ENV; mass_residual", "Currently fixed at 0 under no-leakage assumption", "Should be zero in current model version.", "If positive under no-leakage scenario, closure/design has changed.", out),
        _row("Q_HAUL", r"Q^{HAUL}_t", "Output", "Endogenous flow", "Collection", "tons/year", "t", "eq. 14", "Total formal waste collected and transported by haulers.", "Collection / haulers", "simulate.py", "computed", "model output", "Treatment routing", "Q_HAUL_R + Q_HAUL_O + Q_HAUL_G", "Main formal system throughput.", "Equals total generation under no-leakage assumption.", out),
        _row("Q_HAUL_R", r"Q^{HAUL}_{R,t}", "Output", "Endogenous flow", "Collected recyclables", "tons/year", "t", "eq. 13", "Collected recyclable stream delivered by haulers.", "Collection / haulers", "simulate.py", "computed", "model output", "Q_MRF", "sum_g Q_{g,R,t}", "Higher values increase MRF-bound material before capacity limits.", "Affected by sorting shares and generator scale.", out),
        _row("Q_HAUL_O", r"Q^{HAUL}_{O,t}", "Output", "Endogenous flow", "Collected organics", "tons/year", "t", "eq. 13", "Collected organic stream delivered by haulers.", "Collection / haulers", "simulate.py", "computed", "model output", "Q_CMP", "sum_g Q_{g,O,t}", "Higher values increase composting-bound material before capacity limits.", "Affected by sorting shares and generator scale.", out),
        _row("Q_HAUL_G", r"Q^{HAUL}_{G,t}", "Output", "Endogenous flow", "Collected residual/general", "tons/year", "t", "eq. 13", "Collected general/residual stream delivered by haulers.", "Collection / haulers", "simulate.py", "computed", "model output", "Q_sep; Q_WTE; Q_LF", "sum_g Q_{g,G,t}", "Higher values increase residual treatment burden.", "The main stream affected by source separation policy.", out),
        _row("Q_MRF", r"Q^{MRF}_t", "Output", "Endogenous treatment flow", "MRF throughput", "tons/year", "t", "eq. 31", "Waste processed by the MRF, constrained by available recyclables and MRF capacity.", "Treatment / MRF", "agents/treatment.py", "computed", "model output", "Q_mat; rej_MRF; D_NC", "min(R_avail,K_MRF)", "Higher values increase material recovery up to capacity.", "Overflow is rerouted to residual, not lost.", out),
        _row("Q_CMP", r"Q^{CMP}_t", "Output", "Endogenous treatment flow", "Composting throughput", "tons/year", "t", "eq. 33", "Waste processed by composting/organics treatment, constrained by available organics and capacity.", "Treatment / composting", "agents/treatment.py", "computed", "model output", "Q_cmp; rej_CMP; Q_loss; D_NC", "min(O_avail,K_CMP)", "Higher values increase organics treatment up to capacity.", "Overflow is rerouted to residual, not lost.", out),
        _row("Q_WTE", r"Q^{WTE}_t", "Output", "Endogenous treatment flow", "WTE throughput", "tons/year", "t", "eq. 36", "Residual waste processed by the WTE node, constrained by installed WTE capacity.", "Treatment / WTE", "agents/treatment.py", "computed", "model output", "Q_ash; Q_comb; E_WTE; GHG_dir; D_air", "min(Q_Gavail,K_WTE)", "Higher values reduce direct landfill demand but create ash and emissions.", "Do not interpret WTE as complete waste disappearance; ash remains.", out),
        _row("Q_LF", r"Q^{LF}_t", "Output", "Endogenous treatment flow", "Final landfill demand", "tons/year", "t", "eq. 40", "Final landfill demand including direct residuals, MRF rejects, CMP rejects and WTE ash.", "Treatment / landfill", "agents/treatment.py", "computed", "model output", "S_LF; D_LF; GHG_dir; D_air", "Q_LF_dir + rej_MRF + rej_CMP + Q_ash", "Lower values indicate less final disposal.", "Do not confuse with direct residual before treatment; this includes treatment residuals.", out),
        _row("Q_mat", r"Q^{mat}_t", "Output", "Endogenous recovered output", "Recovered material", "tons/year", "t", "eq. 32", "Recovered marketable material output from MRF throughput.", "Treatment / MRF", "agents/treatment.py", "computed", "model output", "rev_mat; B_env; D_NC", "(1-lambda_MRF)*Q_MRF", "Higher values improve recycling output and external revenue.", "Recovery is net of rejects.", out),
        _row("Q_cmp", r"Q^{cmp}_t", "Output", "Endogenous recovered output", "Compost output", "tons/year", "t", "eq. 34", "Marketable compost output from composting throughput.", "Treatment / composting", "agents/treatment.py", "computed", "model output", "rev_cmp if recorded; B_env; D_NC", "(1-lambda_CMP-ell_CMP)*Q_CMP", "Higher values improve organics recovery.", "Do not include biological loss as recovered compost.", out),
        _row("Q_comb", r"Q^{comb}_t", "Output", "Endogenous residual output", "Combusted mass", "tons/year", "t", "eq. 37 / eq. 55", "WTE input mass converted to flue gas/process outputs excluding ash.", "Treatment / WTE", "agents/treatment.py", "computed", "model output", "mass_residual", "(1-lambda_WTE)*Q_WTE", "Part of mass closure as material leaving solid-waste system boundary.", "Combusted mass is not landfilled mass, but emissions must be counted separately.", out),
        _row("Q_ash", r"Q^{ash}_t", "Output", "Endogenous residual output", "WTE ash", "tons/year", "t", "eq. 37 / eq. 40", "Ash generated by WTE and routed to landfill.", "Treatment / WTE / landfill", "agents/treatment.py", "computed", "model output", "Q_LF", "lambda_WTE * Q_WTE", "Higher values reduce landfill-diversion benefit of WTE.", "Critical for explaining that WTE reduces mass but does not eliminate disposal.", out),
        _row("E_WTE", r"E^{WTE}_t", "Output", "Endogenous energy output", "WTE electricity", "MWh/year", "t", "eq. 38", "Electricity generated by WTE throughput.", "Treatment / WTE", "agents/treatment.py", "computed", "model output", "rev_E; B_env; E_net", "e_E * Q_WTE", "Higher values increase energy revenue and avoided grid emissions.", "Only generated if Q_WTE > 0.", out),
        _row("S_LF", r"S^{LF}_t", "Output", "Endogenous state", "Landfilled stock", "tons", "t", "eq. 41", "Cumulative landfilled stock at the beginning of the year as recorded before annual update.", "Landfill stock", "simulate.py; agents/treatment.py", "computed", "model output", "landfill feasibility; dashboard", "previous S_LF updated by Q_LF", "Tracks cumulative disposal pressure.", "Clarify timing: value recorded before current period update in simulate.py.", out),
        _row("S_ENV", r"S^{ENV}_t", "Output", "Endogenous state", "Informal environmental stock", "tons", "t", "eq. 65", "Informal waste stock in the environment.", "Environment", "simulate.py; agents/environment.py", "computed", "model output", "D_env", "previous S_ENV plus informal", "Should remain zero under no-leakage assumption.", "If no-leakage is assumed, nonzero S_ENV indicates changed model closure.", out),
        _row("gen_to_mun", r"P^{GEN\\to MUN}_t", "Output", "Internal transfer", "Generator fee payments", "$/year", "t", "eq. 9 / eq. 48", "Payments from generators to municipality under stream-differentiated fees.", "Monetary accounting", "simulate.py; agents/generator_base.py", "computed", "model output", "net_MUN", "sum over p_k Q_{g,k}", "Higher values increase municipal revenue.", "Internal transfer; do not count as social cost.", out),
        _row("mun_to_haul", r"P^{MUN\\to HAUL}_t", "Output", "Internal transfer", "Municipal hauler payment", "$/year", "t", "eq. 49", "Municipal contract payment to haulers.", "Monetary accounting", "simulate.py; agents/municipality.py", "computed", "model output", "net_MUN; net_HAUL", "p_HAUL * Q_HAUL", "Higher values increase hauler revenue and municipal expenditure.", "Internal transfer; not automatically a real resource cost.", out),
        _row("tipping", r"R^{tip}_t", "Output", "Internal transfer", "Treatment tipping revenue", "$/year", "t", "eq. 43", "Gate tipping revenue received by treatment operator and paid by haulers.", "Monetary accounting", "simulate.py; agents/treatment.py", "computed", "model output", "net_HAUL; net_T", "t_MRF*Q_MRF + t_CMP*Q_CMP + t_WTE*Q_WTE + t_LF*Q_LF", "Higher values improve treatment operator account and worsen hauler account.", "Internal transfer; distinguish from operating cost.", out),
        _row("rev_mat", r"R^{mat}_t", "Output", "External revenue", "Material sales", "$/year", "t", "eq. 44", "Revenue from selling recovered materials to external buyers.", "Treatment revenues", "simulate.py; agents/treatment.py", "computed", "model output", "net_T; system cost if included", "p_mat * Q_mat", "Higher values improve recycling economics.", "External revenue can be credited in system accounting.", out),
        _row("rev_E", r"R^E_t", "Output", "External revenue", "Electricity sales", "$/year", "t", "eq. 44", "Revenue from selling WTE electricity.", "Treatment revenues", "simulate.py; agents/treatment.py", "computed", "model output", "net_T; system cost if included", "p_E * E_WTE", "Higher values improve WTE economics.", "External revenue can be credited, but should not offset local damages unless using social accounting explicitly.", out),
        _row("landfill_tax", r"\\tau^{LF}_t Q^{LF}_t", "Output", "Internal transfer", "Landfill tax payment", "$/year", "t", "eq. 54", "Landfill tax paid by treatment operator to government/regulator.", "Government / monetary accounting", "simulate.py; agents/government.py", "computed", "model output", "net_T", "tax_LF * Q_LF", "Higher values penalize landfill in operator accounts.", "Transfer if government is inside boundary; avoid double-counting as real resource cost.", out),
        _row("net_MUN", r"\\Pi^{MUN}_t", "Output", "Agent account", "Municipal net balance", "$/year", "t", "budget identity", "Municipal net balance in Stage-0 accounting.", "Monetary accounting", "simulate.py", "computed", "model output", "inspection", "gen_to_mun - mun_to_haul", "Positive values indicate fees exceed hauler payments in current narrow account.", "Not a full municipal budget; education, admin and capital may be excluded depending on implementation.", out),
        _row("net_HAUL", r"\\Pi^{HAUL}_t", "Output", "Agent account", "Hauler net balance", "$/year", "t", "eq. 20 simplified", "Hauler net balance after municipal payment, tipping payments and collection cost.", "Monetary accounting", "simulate.py", "computed", "model output", "inspection", "mun_to_haul - tipping - c_coll*Q_HAUL", "Positive values indicate hauler margin in simplified account.", "Does not necessarily represent real contract profitability unless all hauler costs are included.", out),
        _row("net_T", r"\\Pi^T_t", "Output", "Agent account", "Treatment operator net balance", "$/year", "t", "eq. 47 simplified", "Treatment operator net balance after tipping, product revenues, operating costs and landfill tax.", "Monetary accounting", "simulate.py; agents/treatment.py", "computed", "model output", "inspection", "tipping + rev_mat + rev_cmp + rev_E - C_T_op - landfill_tax", "Positive values indicate operator margin under simplified account.", "Capital, fixed costs and some revenues may be excluded depending on Stage-0 implementation.", out),
        _row("GHG_dir", r"GHG^{dir}_t", "Output", "Environmental output", "Direct GHG", "tCO2e/year", "t", "eq. 66", "Direct GHG emissions from collection, separation and treatment nodes.", "Environmental accounting", "agents/environment.py", "computed", "model output", "GHG_net; D_climate", "emission factors times physical flows", "Higher values worsen climate performance before avoided credits.", "Depends on assumed emission factors, not measured emissions.", out),
        _row("GHG_net", r"GHG^{net}_t", "Output", "Environmental output", "Net GHG", "tCO2e/year", "t", "eq. 70", "Net GHG emissions after subtracting avoided-emissions credits.", "Environmental accounting", "agents/environment.py", "computed", "model output", "D_climate; D_env", "GHG_dir - B_env", "Lower values indicate better climate performance under assumed credits.", "Sensitive to avoided-emissions assumptions; not causal evidence.", out),
        _row("E_net", r"E^{net}_t", "Output", "Environmental output", "Net energy", "MWh/year", "t", "eq. 68", "Net energy balance after WTE generation minus node energy use.", "Environmental accounting", "agents/environment.py", "computed", "model output", "inspection", "e_E*Q_WTE - node energy use", "Higher values indicate more net electricity output.", "Requires consistent energy-use assumptions.", out),
        _row("D_climate", r"D^{climate}_t", "Output", "Monetized damage", "Climate damage", "$/year", "t", "eq. 71", "Monetized climate damage from net GHG emissions.", "Environmental damages", "agents/environment.py", "computed", "model output", "D_env", "SCC * GHG_net", "Higher values worsen social cost.", "Sensitive to SCC and avoided-emissions assumptions.", out),
        _row("D_air", r"D^{air}_t", "Output", "Monetized damage", "Local air damage", "$/year", "t", "eq. 72", "Monetized local air-pollution damage from pollutant emissions.", "Environmental damages", "agents/environment.py", "computed", "model output", "D_env", "sum_p D_p * AP_p", "Higher values worsen social cost.", "Highly sensitive to emission and damage factors; spatial exposure is not represented in Stage-0.", out),
        _row("D_env", r"D^{env}_t", "Output", "Monetized damage", "Total environmental damage", "$/year", "t", "eq. 73", "Total monetized environmental damage from climate, local air pollution and informal stock damages.", "Environmental damages", "agents/environment.py", "computed", "model output", "system social cost", "D_climate + D_air + d_ENV*S_ENV", "Higher values worsen social performance.", "Not a complete welfare measure; only includes modeled externalities.", out),
        _row("D_NC", r"D^{NC}_t", "Output", "Indicator", "Non-combustion diversion", "share", "t", "eq. 56", "Non-combustion diversion rate based on recovered material and compost outputs over formal collection.", "Closure / indicators", "closure.py", "computed", "model output", "dashboard; policy targets", "net recovered non-combustion tonnage / Q_HAUL", "Higher values indicate more recycling/composting diversion.", "Does not credit WTE; central distinction for policy interpretation.", out),
        _row("D_LF", r"D^{LF}_t", "Output", "Indicator", "Landfill diversion", "share", "t", "eq. 57", "Landfill diversion rate counting all pathways avoiding final landfill demand, including WTE.", "Closure / indicators", "closure.py", "computed", "model output", "dashboard; policy targets", "1 - Q_LF / Q_HAUL", "Higher values indicate less final landfill demand.", "Can rise with WTE even if non-combustion diversion does not improve.", out),
        _row("mass_residual", "", "Output", "Diagnostic", "Mass closure", "tons/year", "t", "eq. 55", "Signed residual of system-wide physical mass balance.", "Closure / diagnostics", "closure.py", "computed", "model output", "Closure_Checks", "generation - accounted mass", "Should be approximately zero in every year.", "If nonzero beyond tolerance, model accounting is invalid for that run.", out),
    ]

    return rows


def build_all_variables(
    df: pd.DataFrame,
    exogenous_df: pd.DataFrame | None = None,
    params_df: pd.DataFrame | None = None,
    air_df: pd.DataFrame | None = None,
) -> pd.DataFrame:
    """Build a single master dictionary for inputs and outputs."""
    rows = _manual_variable_catalogue()
    dictionary = pd.DataFrame(rows, columns=MASTER_COLUMNS)

    documented = set(dictionary["variable"])
    extra_rows: list[dict] = []

    def append_unknown(var: str, where: str, input_output: str):
        if var not in documented:
            extra_rows.append(_row(
                variable=var,
                input_output=input_output,
                type="Undocumented variable",
                subtype="Needs classification",
                definition="Automatically detected but not yet documented in the manual catalogue.",
                source_status="unknown",
                source=where,
                critical_note="Add equation, definition, unit and interpretation before using this variable in a paper or policy claim.",
                appears_in_sheet=where,
            ))
            documented.add(var)

    if params_df is not None and "name" in params_df.columns:
        for var in params_df["name"].dropna().astype(str):
            append_unknown(var, "Parameters_Assumptions", "Input")

    if exogenous_df is not None:
        for var in exogenous_df.columns.astype(str):
            append_unknown(var, "Exogenous_Inputs", "Input")

    if air_df is not None:
        for var in air_df.columns.astype(str):
            append_unknown(var, "Air_Pollutant_Factors", "Input")

    for var in df.columns.astype(str):
        append_unknown(var, "Endogenous_Outputs", "Output")

    if extra_rows:
        dictionary = pd.concat([dictionary, pd.DataFrame(extra_rows)], ignore_index=True)

    dictionary = dictionary.sort_values(
        by=["input_output", "type", "model_block", "variable"],
        ascending=[True, True, True, True],
        kind="stable",
    ).reset_index(drop=True)
    return dictionary[MASTER_COLUMNS]


# Backward-compatible alias used by older code/tests.
def build_variable_dictionary(df: pd.DataFrame) -> pd.DataFrame:
    return build_all_variables(df)


# Bloque 4 — additional workbook tables: flow map, summaries and checks

def build_flow_map() -> pd.DataFrame:
    rows = [
        {
            "Block": "Generator behavior",
            "Inputs": "N, gbar_g, rho_min_g, kappa_rho, A, Psi, p_G, p_R, p_O",
            "Equations": "eqs. 3, 8, 9",
            "Outputs": "gen_total, Q_HAUL_R, Q_HAUL_O, Q_HAUL_G, gen_to_mun",
            "Main risk / limitation": "Generation and sorting responses are assumed closed forms, not estimated behavioral relationships.",
        },
        {
            "Block": "Municipality",
            "Inputs": "E_edu, X_MUN, gamma, delta_psi, p_HAUL",
            "Equations": "eqs. 10a, 10b, 49, 51",
            "Outputs": "A, Psi, mun_to_haul, net_MUN",
            "Main risk / limitation": "Education and convenience channels depend on calibration of stock-transition parameters.",
        },
        {
            "Block": "Haulers",
            "Inputs": "Q_HAUL, p_HAUL, tipping fees, c_coll",
            "Equations": "eqs. 13-14, 20 simplified",
            "Outputs": "Q_HAUL, net_HAUL",
            "Main risk / limitation": "Collection cost is a simple per-ton coefficient; route distance and spatial heterogeneity are not explicit.",
        },
        {
            "Block": "Treatment",
            "Inputs": "Q_HAUL_R, Q_HAUL_O, Q_HAUL_G, sep_share, eta_R, eta_O, K_MRF, K_CMP, K_WTE, lambda parameters",
            "Equations": "eqs. 22-40, 43-46",
            "Outputs": "Q_MRF, Q_CMP, Q_WTE, Q_LF, Q_mat, Q_cmp, Q_ash, E_WTE, net_T",
            "Main risk / limitation": "Routing is deterministic and capacity-constrained; treatment operator optimization is not solved in Stage-0.",
        },
        {
            "Block": "Environment",
            "Inputs": "Treatment flows, GHG factors, air pollutant factors, SCC, D_p",
            "Equations": "eqs. 65-73",
            "Outputs": "GHG_dir, GHG_net, E_net, D_climate, D_air, D_env",
            "Main risk / limitation": "Air damages are not spatially explicit; emission and damage factors require strong documentation.",
        },
        {
            "Block": "Closure",
            "Inputs": "All physical flows and treatment residuals",
            "Equations": "eqs. 55-57",
            "Outputs": "mass_residual, D_NC, D_LF",
            "Main risk / limitation": "Closure validates accounting, not empirical realism or causal identification.",
        },
    ]
    return pd.DataFrame(rows)


def build_classification_summary(all_variables: pd.DataFrame) -> pd.DataFrame:
    summary = (
        all_variables
        .groupby(["input_output", "type"], as_index=False)
        .size()
        .rename(columns={"size": "count"})
    )
    return summary


def build_closure_checks(df: pd.DataFrame) -> pd.DataFrame:
    checks = pd.DataFrame()
    if "year" in df.columns:
        checks["year"] = df["year"]
    if "mass_residual" in df.columns:
        checks["mass_residual"] = df["mass_residual"]
        checks["mass_status"] = df["mass_residual"].abs().apply(lambda x: "OK" if x < 1e-3 else "CHECK")
    if "gen_total" in df.columns:
        checks["gen_total"] = df["gen_total"]
    if "Q_HAUL" in df.columns:
        checks["Q_HAUL"] = df["Q_HAUL"]
    if "Q_LF" in df.columns:
        checks["Q_LF"] = df["Q_LF"]
    if "D_NC" in df.columns:
        checks["D_NC"] = df["D_NC"]
    if "D_LF" in df.columns:
        checks["D_LF"] = df["D_LF"]
    if "S_LF" in df.columns:
        checks["S_LF"] = df["S_LF"]
    if "S_ENV" in df.columns:
        checks["S_ENV"] = df["S_ENV"]
        checks["leakage_status"] = df["S_ENV"].apply(lambda x: "OK" if abs(x) < 1e-9 else "CHECK")
    return checks


# Bloque 5 — input-data documentation builders
def build_input_data_index(
    exogenous_file: Path,
    params_file: Path,
    air_file: Path,
    initial_file: Path | None = None,
) -> pd.DataFrame:
    """Index of input files included in the documentation workbook."""
    rows = [
        {
            "input_file": str(params_file),
            "sheet_in_report": "Parameters_Assumptions",
            "model_role": "Time-invariant scalar parameters and model switches.",
            "required": True,
            "critical_note": (
                "These are assumptions/calibrations. They should not be interpreted "
                "as observed facts unless source_status says so."
            ),
        },
        {
            "input_file": str(exogenous_file),
            "sheet_in_report": "Exogenous_Inputs",
            "model_role": "Annual exogenous drivers, market prices, policy variables and capacities.",
            "required": True,
            "critical_note": (
                "These define the scenario path. Changes here can mechanically alter "
                "results without changing model structure."
            ),
        },
        {
            "input_file": str(air_file),
            "sheet_in_report": "Air_Pollutant_Factors",
            "model_role": "Local air-pollutant emission factors and marginal damage values.",
            "required": True,
            "critical_note": (
                "High uncertainty. Facility-specific emissions and spatial damage estimates "
                "are needed before strong policy claims."
            ),
        },
    ]

    if initial_file is not None:
        rows.append({
            "input_file": str(initial_file),
            "sheet_in_report": "Initial_Conditions",
            "model_role": "Initial state values, if stored separately.",
            "required": False,
            "critical_note": (
                "In the current implementation, several initial states are stored "
                "in params_scalar.csv instead."
            ),
        })

    return pd.DataFrame(rows)


def build_documented_exogenous_inputs(
    exogenous_df: pd.DataFrame,
    all_variables: pd.DataFrame,
) -> pd.DataFrame:
    """
    Build a documented version of timeseries_exog.csv.

    Each row is one annual exogenous/scenario variable. The sheet includes:
    - the variable definition from ALL_VARIABLES
    - unit, equation, source status and critical note
    - the actual annual values used in the run

    This replaces the earlier Exogenous_Inputs_Transposed sheet.
    """
    if exogenous_df is None or exogenous_df.empty:
        return pd.DataFrame()

    if "year" not in exogenous_df.columns:
        return pd.DataFrame({
            "warning": ["Cannot document exogenous inputs because column 'year' is missing."]
        })

    documentation_cols = [
        "variable",
        "symbol",
        "input_output",
        "type",
        "subtype",
        "unit",
        "time_index",
        "equation",
        "definition",
        "model_block",
        "code_module",
        "source_status",
        "source",
        "interpretation",
        "critical_note",
    ]

    doc = all_variables[documentation_cols].copy()

    year_labels = [
        str(int(y)) if isinstance(y, (int, float)) and float(y).is_integer() else str(y)
        for y in exogenous_df["year"].tolist()
    ]

    rows = []

    for variable in exogenous_df.columns:
        if variable == "year":
            continue

        row = {"variable": variable}
        values = exogenous_df[variable].tolist()

        for year_label, value in zip(year_labels, values):
            row[year_label] = value

        rows.append(row)

    values_df = pd.DataFrame(rows)
    documented = values_df.merge(doc, on="variable", how="left")

    missing = documented["definition"].isna()
    if missing.any():
        documented.loc[missing, "definition"] = (
            "Exogenous input detected in timeseries_exog.csv but not yet documented "
            "in ALL_VARIABLES."
        )
        documented.loc[missing, "critical_note"] = (
            "Add definition, unit, equation, source status and interpretation before "
            "using this input in a paper or policy claim."
        )

    doc_order = [
        "variable",
        "symbol",
        "input_output",
        "type",
        "subtype",
        "unit",
        "time_index",
        "equation",
        "definition",
        "model_block",
        "code_module",
        "source_status",
        "source",
        "interpretation",
        "critical_note",
    ]

    value_cols = [c for c in documented.columns if c not in doc_order]

    return documented[doc_order + value_cols]


def build_documented_inputs(
    input_df: pd.DataFrame,
    all_variables: pd.DataFrame,
    variable_col: str,
) -> pd.DataFrame:
    """
    Merge raw input values with the master variable dictionary.

    Used for scalar parameters and similar input tables where each row contains
    one variable name.
    """
    if input_df is None or input_df.empty:
        return pd.DataFrame()

    if variable_col not in input_df.columns:
        return pd.DataFrame({
            "warning": [f"Cannot document input sheet because column '{variable_col}' is missing."]
        })

    documentation_cols = [
        "variable",
        "symbol",
        "input_output",
        "type",
        "subtype",
        "unit",
        "time_index",
        "equation",
        "definition",
        "model_block",
        "code_module",
        "source_status",
        "source",
        "interpretation",
        "critical_note",
    ]

    doc = all_variables[documentation_cols].copy()

    left = input_df.copy()
    left = left.rename(columns={variable_col: "variable"})

    merged = left.merge(doc, on="variable", how="left")

    missing = merged["definition"].isna()
    if missing.any():
        merged.loc[missing, "definition"] = (
            "Input variable detected in CSV but not yet documented in ALL_VARIABLES."
        )
        merged.loc[missing, "critical_note"] = (
            "Add definition, unit, equation, source status and interpretation before "
            "using this input in a paper or policy claim."
        )

    return merged


def build_air_factor_dictionary() -> pd.DataFrame:
    """Definitions of the columns used in air_pollutant_factors.csv."""
    return pd.DataFrame([
        {
            "column": "pollutant",
            "definition": "Name of the local air pollutant being tracked.",
            "unit": "text",
            "interpretation": "Identifies pollutant p in the local air-pollution equation.",
            "equation": "eq. 67 / eq. 72",
            "critical_note": (
                "Use code-safe labels consistently. PM25 may represent PM2.5 "
                "if dots are inconvenient in code."
            ),
        },
        {
            "column": "a_WTE",
            "definition": "Emission factor of pollutant p from Waste-to-Energy treatment.",
            "unit": "kg pollutant / ton WTE throughput",
            "interpretation": "How many kg of pollutant p are emitted per ton sent to WTE.",
            "equation": "eq. 67",
            "critical_note": "Requires facility-specific or literature-supported emissions evidence.",
        },
        {
            "column": "a_LF",
            "definition": "Emission factor of pollutant p from landfill disposal.",
            "unit": "kg pollutant / ton landfilled",
            "interpretation": "How many kg of pollutant p are emitted per ton finally sent to landfill.",
            "equation": "eq. 67",
            "critical_note": (
                "Should reflect whether the pollutant is actually relevant for landfill emissions."
            ),
        },
        {
            "column": "a_sep",
            "definition": "Emission factor of pollutant p from mechanical separation.",
            "unit": "kg pollutant / ton separated",
            "interpretation": (
                "How many kg of pollutant p are emitted per ton processed through "
                "the separation node."
            ),
            "equation": "eq. 67",
            "critical_note": "Usually small but relevant when sep_share is high.",
        },
        {
            "column": "D_p",
            "definition": "Marginal damage of pollutant p.",
            "unit": "$/kg pollutant emitted",
            "interpretation": "Converts physical local-air emissions into monetized damages.",
            "equation": "eq. 72",
            "critical_note": (
                "Highly spatial and exposure-dependent. Placeholder values should not "
                "support strong policy conclusions."
            ),
        },
    ])

# Bloque 6 — workbook README
def build_readme() -> pd.DataFrame:
    return pd.DataFrame([
        {
            "sheet": "README",
            "purpose": "Explains the workbook structure.",
            "main_risk": "None.",
        },
        {
            "sheet": "Input_Data_Index",
            "purpose": "Lists the exact input files included in the report.",
            "main_risk": "If this sheet is missing, the workbook is not auditable.",
        },
        {
            "sheet": "ALL_VARIABLES",
            "purpose": "Single master dictionary for every input and output variable.",
            "main_risk": "This is the authoritative documentation sheet; keep it complete.",
        },
        {
            "sheet": "Classification_Summary",
            "purpose": "Counts variables by input/output and type.",
            "main_risk": "Useful for detecting undocumented variables.",
        },
        {
            "sheet": "Flow_Map",
            "purpose": "Shows how model blocks transform inputs into outputs.",
            "main_risk": "This is conceptual, not a causal identification map.",
        },
        {
            "sheet": "Closure_Checks",
            "purpose": "Shows mass-balance diagnostics and key closure indicators by year.",
            "main_risk": "Closure validates accounting only; it does not validate parameter realism.",
        },
        {
            "sheet": "Endogenous_Outputs",
            "purpose": "Raw outputs produced by simulate.py.",
            "main_risk": "Outputs are mechanical results under assumptions, not empirical causal estimates.",
        },
        {
            "sheet": "Exogenous_Inputs",
            "purpose": "Annual scenario variables read before simulation, in machine-readable format.",
            "main_risk": "Scenario assumptions must be distinguished from observed data.",
        },
        {
            "sheet": "Exogenous_Inputs_Documented",
            "purpose": (
                "Annual exogenous/scenario variables merged with definitions, units, equations, "
                "source status and critical notes from ALL_VARIABLES."
            ),
            "main_risk": (
                "This is the main sheet for checking whether scenario inputs are conceptually documented."
            ),
        },
        {
            "sheet": "Parameters_Assumptions",
            "purpose": "Scalar parameters and assumptions read before simulation.",
            "main_risk": "Behavioral and environmental parameters need calibration/sensitivity analysis.",
        },
        {
            "sheet": "Parameters_Documented",
            "purpose": "Scalar parameters merged with definitions from ALL_VARIABLES.",
            "main_risk": "Use this to inspect whether each parameter is active, assumed, calibrated or legacy.",
        },
        {
            "sheet": "Air_Pollutant_Factors",
            "purpose": "Local air-pollutant emission factors and marginal damages.",
            "main_risk": (
                "Highly sensitive; should use facility-specific or damage-model evidence before strong claims."
            ),
        },
        {
            "sheet": "Air_Factors_Documented",
            "purpose": "Air-pollutant input table merged with variable definitions.",
            "main_risk": "Do not interpret placeholder pollutant factors as facility-specific evidence.",
        },
        {
            "sheet": "Air_Factor_Dictionary",
            "purpose": "Defines pollutant, a_WTE, a_LF, a_sep and D_p.",
            "main_risk": "Essential because these column names are not self-explanatory.",
        },
        {
            "sheet": "Initial_Conditions",
            "purpose": "Optional sheet with initial state values if a separate initial_conditions.csv exists.",
            "main_risk": "May be absent if initial states are stored in params_scalar.csv.",
        },
    ])


# Bloque 7 — Excel writer utilities

def _write_sheet(writer, df: pd.DataFrame, sheet_name: str) -> None:
    df.to_excel(writer, sheet_name=sheet_name, index=False)


def _format_workbook(writer) -> None:
    """Basic readable formatting using openpyxl through pandas' writer."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.formatting.rule import CellIsRule

    workbook = writer.book
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(bold=True, color="FFFFFF")
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    text_heavy = {
        "definition", "critical_note", "interpretation", "used_in", "computed_from",
        "source", "code_module", "main_risk", "purpose", "Main risk / limitation",
        "Inputs", "Outputs", "Equations",
    }

    for sheet_name in workbook.sheetnames:
        ws = workbook[sheet_name]
        ws.freeze_panes = "A2"
        if ws.max_row > 1 and ws.max_column > 1:
            ws.auto_filter.ref = ws.dimensions

        # Header style
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border
        ws.row_dimensions[1].height = 28

        # Determine widths from header names, capped to avoid monster cells.
        for col in ws.iter_cols(min_row=1, max_row=min(ws.max_row, 50)):
            col_letter = col[0].column_letter
            header = str(col[0].value or "")
            max_len = max(len(str(c.value)) if c.value is not None else 0 for c in col)
            if header in text_heavy:
                width = min(max(max_len + 2, 24), 48)
            elif sheet_name == "Endogenous_Outputs":
                width = min(max(max_len + 2, 10), 16)
            else:
                width = min(max(max_len + 2, 10), 28)
            ws.column_dimensions[col_letter].width = width

        # Body style
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.border = border
        for r in range(2, ws.max_row + 1):
            ws.row_dimensions[r].height = 36 if sheet_name == "ALL_VARIABLES" else 22

        # Number formats and conditional flags
        header_map = {str(ws.cell(1, c).value): c for c in range(1, ws.max_column + 1)}
        for name in ["mass_residual", "gen_total", "Q_HAUL", "Q_LF", "S_LF", "S_ENV"]:
            if name in header_map:
                col_letter = ws.cell(1, header_map[name]).column_letter
                for cell in ws[f"{col_letter}2:{col_letter}{ws.max_row}"]:
                    cell[0].number_format = "#,##0.000"
        for name in ["D_NC", "D_LF"]:
            if name in header_map:
                col_letter = ws.cell(1, header_map[name]).column_letter
                for cell in ws[f"{col_letter}2:{col_letter}{ws.max_row}"]:
                    cell[0].number_format = "0.0%"

        # Highlight CHECK statuses if present.
        for status_col in ["mass_status", "leakage_status"]:
            if status_col in header_map and ws.max_row >= 2:
                col_letter = ws.cell(1, header_map[status_col]).column_letter
                rng = f"{col_letter}2:{col_letter}{ws.max_row}"
                ws.conditional_formatting.add(
                    rng,
                    CellIsRule(operator="equal", formula=['"CHECK"'],
                               fill=PatternFill("solid", fgColor="F4CCCC"))
                )


# Bloque 8 — main report writer
def write_model_excel_report(
    df: pd.DataFrame,
    out_path: str | Path = "MSW_Model_Documentation.xlsx",
    exogenous_path: str | Path | None = None,
    params_path: str | Path | None = None,
    air_pollutants_path: str | Path | None = None,
    initial_conditions_path: str | Path | None = None,
) -> None:
    """
    Writes an Excel workbook documenting the Stage-0 MSW model.

    Central design principle:
    - ALL_VARIABLES is the single master dictionary for both inputs and outputs.
    - Input sheets contain the exact CSV input values used by the run.
    - Output sheets contain computed results.
    - Closure_Checks validates accounting consistency, not empirical realism.

    The function fails clearly if required input CSVs are missing. A documentation
    workbook without inputs is not auditable.
    """

    out_path = Path(out_path)

    # Resolve required input files.
    # Prefer explicit paths if passed from run.py; otherwise use package-level DATA_DIR.
    exogenous_file = _required_input_path(
        "timeseries_exog.csv",
        exogenous_path,
        DEFAULT_EXOGENOUS_PATH,
        "data/timeseries_exog.csv",
    )
    params_file = _required_input_path(
        "params_scalar.csv",
        params_path,
        DEFAULT_PARAMS_PATH,
        "data/params_scalar.csv",
        "data/params.csv",
    )
    air_file = _required_input_path(
        "air_pollutant_factors.csv",
        air_pollutants_path,
        DEFAULT_AIR_PATH,
        "data/air_pollutant_factors.csv",
    )

    # Optional initial-conditions file.
    initial_file = _existing_path(
        initial_conditions_path,
        DEFAULT_INITIAL_PATH,
        "data/initial_conditions.csv",
    )

    # Read inputs.
    exogenous_df = _read_csv_auto(exogenous_file)
    params_df = _read_csv_auto(params_file)
    air_df = _read_csv_auto(air_file)
    initial_df = _read_csv_auto(initial_file) if initial_file else None

    # Build documentation tables.
    all_variables = build_all_variables(
        df=df,
        exogenous_df=exogenous_df,
        params_df=params_df,
        air_df=air_df,
    )

    input_data_index = build_input_data_index(
        exogenous_file=exogenous_file,
        params_file=params_file,
        air_file=air_file,
        initial_file=initial_file,
    )

    classification_summary = build_classification_summary(all_variables)
    flow_map = build_flow_map()
    closure_checks = build_closure_checks(df)
    readme = build_readme()

    documented_exogenous = build_documented_exogenous_inputs(
        exogenous_df=exogenous_df,
        all_variables=all_variables,
    )

    documented_params = build_documented_inputs(
        input_df=params_df,
        all_variables=all_variables,
        variable_col="name",
    )

    documented_air = build_documented_inputs(
        input_df=air_df,
        all_variables=all_variables,
        variable_col="pollutant",
    )

    air_factor_dictionary = build_air_factor_dictionary()

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        # Overview and documentation
        _write_sheet(writer, readme, "README")
        _write_sheet(writer, input_data_index, "Input_Data_Index")
        _write_sheet(writer, all_variables, "ALL_VARIABLES")
        _write_sheet(writer, classification_summary, "Classification_Summary")
        _write_sheet(writer, flow_map, "Flow_Map")
        _write_sheet(writer, closure_checks, "Closure_Checks")

        # Outputs
        _write_sheet(writer, df, "Endogenous_Outputs")

        # Raw inputs
        _write_sheet(writer, exogenous_df, "Exogenous_Inputs")
        _write_sheet(writer, params_df, "Parameters_Assumptions")
        _write_sheet(writer, air_df, "Air_Pollutant_Factors")

        # Documented inputs
        _write_sheet(writer, documented_exogenous, "Exogenous_Inputs_Documented")
        _write_sheet(writer, documented_params, "Parameters_Documented")
        _write_sheet(writer, documented_air, "Air_Factors_Documented")
        _write_sheet(writer, air_factor_dictionary, "Air_Factor_Dictionary")

        if initial_df is not None:
            _write_sheet(writer, initial_df, "Initial_Conditions")

        _format_workbook(writer)

    print(f"Excel report written to: {out_path}")
